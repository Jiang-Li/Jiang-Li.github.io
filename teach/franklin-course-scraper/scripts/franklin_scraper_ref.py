#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Franklin University Course Scraper - Production Version

A comprehensive web scraper for Franklin University's course search system.
Extracts detailed course information including schedules, enrollment data,
instructor details, and teaching modes. Features intelligent Excel output
with conditional formatting for change tracking.

Key Features:
- Automated course data extraction via Selenium WebDriver
- Professional Excel output with table formatting
- Real-time conditional formatting for change tracking
- Robust error handling and performance optimizations
- Term filtering for focused data collection
- Preservation of user data during updates

Author: Course Category Project
Version: 4.0 (Production Enhanced)
Last Updated: May 2025
"""

import time
import os
import re
from datetime import datetime
from dataclasses import dataclass
from typing import List, Optional, Tuple
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule

# ==============================================================================
# CONFIGURATION CONSTANTS
# ==============================================================================

# Weekday mapping - Comprehensive handling for various university formats
# Handles standard abbreviations and Franklin University's specific Thursday formats (R, Th, TH)
WEEKDAY_MAPPING = {
    'M': 'Monday',
    'T': 'Tuesday', 
    'W': 'Wednesday',
    'R': 'Thursday',  # Franklin University format for Thursday
    'Th': 'Thursday',  # Alternative Thursday format
    'TH': 'Thursday',  # Uppercase Thursday format
    'F': 'Friday',
    'S': 'Saturday',
    'U': 'Sunday'  # Franklin University format for Sunday
}

# Short weekday names for compact display
WEEKDAY_MAPPING_SHORT = {
    'M': 'Mon',
    'T': 'Tue', 
    'W': 'Wed',
    'R': 'Thu',  # Franklin University format
    'Th': 'Thu',  # Alternative format
    'TH': 'Thu',  # Uppercase format
    'F': 'Fri',
    'S': 'Sat',
    'U': 'Sun'
}

# ==============================================================================
# DATA MODELS
# ==============================================================================

@dataclass
class CourseRequest:
    """
    Configuration data model for course scraping requests.
    
    Attributes:
        term (str): Academic term filter (e.g., "Spring 2025", empty string for all terms)
        courses (List[Tuple[str, bool]]): List of (course_code, is_first_term) tuples
                                        where is_first_term indicates special first-term handling
    """
    term: str
    courses: List[Tuple[str, bool]]  # (course_code, is_first_term)

@dataclass
class CourseSection:
    """
    Comprehensive data model for individual course sections.
    
    Contains all extracted information for a single course section including
    enrollment data, scheduling information, instructor details, and metadata.
    """
    # Core identification
    course_code: str      # e.g., "DATA*610" 
    session_code: str     # e.g., "Q1FF"
    course_name: str      # Full course title
    credits: str          # Credit hours
    
    # Enrollment information  
    seats_available: str  # Available seats count
    seats_total: str      # Total capacity
    seats_waitlisted: str # Waitlist count
    
    # Schedule details
    weekdays: List[str]      # Full weekday names ["Monday", "Wednesday"]
    weekdays_short: List[str] # Short names ["Mon", "Wed"] 
    class_times: List[str]   # Time ranges ["6:00 PM - 8:00 PM"]
    
    # Location and faculty
    locations: List[str]     # Classroom/building information
    instructors: List[str]   # Faculty names
    teaching_mode: str       # Delivery method (Face-to-Face, Online, etc.)
    
    # Academic calendar
    start_date: str         # Course start date
    end_date: str           # Course end date
    term: str               # Academic term
    
    # Additional metadata
    prerequisites: str = ""      # Course prerequisites
    course_description: str = "" # Course description
    is_first_term: bool = False  # Special first term indicator

# ==============================================================================
# MAIN SCRAPER CLASS
# ==============================================================================

class FranklinCourseScraper:
    """
    Main scraper class for Franklin University course information extraction.
    
    This class handles all aspects of web scraping including browser automation,
    data extraction, Excel output generation, and error handling. Designed for
    production use with comprehensive logging and performance optimizations.
    """
    
    def __init__(self, headless=True):
        """
        Initialize the course scraper with browser configuration.
        
        Args:
            headless (bool): Run browser in headless mode for better performance
        """
        self.base_url = "https://selfservice.franklin.edu/Student/Courses/Search"
        self.driver = None
        self.headless = headless
        self.setup_driver()
    
    def setup_driver(self):
        """
        Configure and initialize Chrome WebDriver with performance optimizations.
        
        Sets up Chrome with various flags for optimal scraping performance:
        - Disables unnecessary features (images, plugins, etc.)
        - Configures user agent and window size
        - Sets page load strategy for faster navigation
        
        Raises:
            Exception: If driver initialization fails
        """
        chrome_options = Options()
        if self.headless:
            chrome_options.add_argument("--headless")
        
        # Core performance optimizations - reduce resource usage
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--disable-web-security")
        chrome_options.add_argument("--disable-features=VizDisplayCompositor")
        chrome_options.add_argument("--disable-extensions")
        chrome_options.add_argument("--disable-plugins")
        chrome_options.add_argument("--disable-images")  # Critical: Disable image loading for speed
        chrome_options.add_argument("--disable-javascript-harmony-shipping")
        chrome_options.add_argument("--disable-background-timer-throttling")
        chrome_options.add_argument("--disable-renderer-backgrounding")
        chrome_options.add_argument("--disable-backgrounding-occluded-windows")
        
        # Browser configuration
        chrome_options.add_argument("--window-size=1920,1080")
        chrome_options.add_argument("--user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")
        
        # Faster page loading strategy
        chrome_options.add_argument("--page-load-strategy=eager")
        
        try:
            self.driver = webdriver.Chrome(options=chrome_options)
            # Reduce implicit wait for faster element detection
            self.driver.implicitly_wait(1)
            print("✅ Chrome driver initialized successfully")
        except Exception as e:
            print(f"❌ Driver initialization failed: {e}")
            raise
    
    def read_course_list(self, filename: str = "course_request.md") -> CourseRequest:
        """Read course request from course_request.md file with simplified format"""
        try:
            if not os.path.exists(filename):
                print(f"⚠️  {filename} file not found, using default configuration")
                return CourseRequest("", [("DATA 610", False)])
            
            with open(filename, 'r', encoding='utf-8') as f:
                content = f.read()
            
            # Find the Current Configuration section and its code block
            config_match = re.search(r'## Current Configuration\s*```\n(.*?)\n```', content, re.DOTALL)
            if not config_match:
                print(f"⚠️  No Current Configuration section found in {filename}, using default")
                return CourseRequest("", [("DATA 610", False)])
            
            # Process the configuration block
            config_block = config_match.group(1)
            lines = [line.strip() for line in config_block.split('\n') if line.strip()]
            
            if not lines:
                print(f"⚠️  Empty configuration block in {filename}, using default")
                return CourseRequest("", [("DATA 610", False)])
            
            term = None
            courses = []
            
            for line in lines:
                # Skip comments
                if line.startswith('#') or line.startswith('//'):
                    continue
                
                # Check for term line: "Term: Spring 2025"
                if line.startswith('Term:'):
                    term = line.replace('Term:', '').strip()
                    continue
                
                # Course line: "*DATA 610" or "DATA 610"
                if re.match(r'^[*]?[A-Za-z]+\s+\d+', line):
                    # Remove any trailing comments
                    line = line.split('#')[0].strip()
                    is_first_term = line.startswith('*')
                    course_code = line[1:].strip() if is_first_term else line.strip()
                    courses.append((course_code, is_first_term))
            
            # If we found courses, return the request
            if courses:
                print(f"📋 Configuration found:")
                if term:
                    print(f"   Term: {term}")
                else:
                    print(f"   Term: All terms (no filter)")
                    term = ""  # Empty string means no term filtering
                print(f"   Courses: {len(courses)}")
                for course, is_first in courses:
                    first_term_indicator = "*" if is_first else " "
                    print(f"     {first_term_indicator} {course}")
                return CourseRequest(term, courses)
            
            print(f"⚠️  No valid configuration found in {filename}, using default")
            return CourseRequest("", [("DATA 610", False)])
            
        except Exception as e:
            print(f"❌ Failed to read {filename}: {e}")
            return CourseRequest("", [("DATA 610", False)])

    def search_course(self, course_code: str, term: str) -> bool:
        """Search for course with term filtering"""
        try:
            # Format course code for URL - replace space with + but preserve *
            search_code = course_code.replace(' ', '+')
            if '*' not in search_code:  # Only add * if not already present
                parts = search_code.split('+')
                if len(parts) == 2:
                    search_code = f"{parts[0]}*{parts[1]}"
            
            # Construct search URL with term filtering
            search_url = f"{self.base_url}?keyword={search_code}"
            
            # Add term filter if specified
            if term:
                # Convert term to URL parameter format
                term_param = term.replace(' ', '+')
                search_url += f"&term={term_param}"
                print(f"🔍 Searching: {course_code} (Term: {term})")
            else:
                print(f"🔍 Searching: {course_code} (All terms)")
            
            self.driver.get(search_url)
            
            # First try smart waiting for specific elements (fast path)
            try:
                WebDriverWait(self.driver, 2).until(
                    lambda driver: (
                        len(driver.find_elements(By.XPATH, "//a[contains(text(), 'View Available Sections')]")) > 0 or
                        len(driver.find_elements(By.XPATH, "//div[contains(@class, 'no-results')]")) > 0 or
                        len(driver.find_elements(By.XPATH, "//table[contains(@class, 'search-sectiontable')]")) > 0
                    )
                )
                return True
            except TimeoutException:
                # Fallback: wait for page to be generally ready
                try:
                    WebDriverWait(self.driver, 3).until(
                        EC.presence_of_element_located((By.TAG_NAME, "body"))
                    )
                    # Additional check for page content
                    WebDriverWait(self.driver, 2).until(
                        lambda driver: driver.execute_script("return document.readyState") == "complete"
                    )
                    return True
                except TimeoutException:
                    print("❌ Search page loading timeout")
                    return False
            
        except Exception as e:
            print(f"❌ Error during search: {e}")
            return False
    
    def click_view_sections(self, course_code: str) -> bool:
        """Click view available sections link"""
        try:
            # Smart wait - check if sections are already visible or if we need to click a link
            try:
                # First check if course sections are already visible (very fast check)
                if len(self.driver.find_elements(By.CLASS_NAME, "search-sectiontable")) > 0:
                    return True
            except:
                pass
            
            # Format course code for link matching - ensure proper asterisk handling
            search_code = course_code.replace(' ', '-')
            if '*' not in search_code:
                parts = search_code.split('-')
                if len(parts) == 2:
                    search_code = f"{parts[0]}*{parts[1]}"
            
            # Look for "View Available Sections" link with optimized waiting
            selectors = [
                f"//a[contains(text(), 'View Available Sections for {search_code.upper()}')]",
                f"//a[contains(text(), 'View Available Sections for {course_code.upper().replace(' ', '-')}')]",
                f"//a[contains(text(), 'View Available Sections')]",
                "//a[contains(@href, 'sections') or contains(text(), 'sections')]",
                "//button[contains(text(), 'View Available Sections')]",
                "//*[contains(text(), 'View Available Sections')]"
            ]
            
            link_element = None
            for selector in selectors:
                try:
                    # Faster element detection with shorter timeout
                    elements = self.driver.find_elements(By.XPATH, selector)
                    if elements:
                        link_element = elements[0]
                        # Check if element is immediately clickable
                        if link_element.is_displayed() and link_element.is_enabled():
                            break
                except:
                    continue
            
            if not link_element:
                return True
            
            # Click immediately without scrolling or additional waits
            try:
                link_element.click()
            except:
                # Fallback to JavaScript click if regular click fails
                self.driver.execute_script("arguments[0].click();", link_element)
            
            # Extended wait for content update with more debugging
            start_wait = time.time()
            max_wait = 5.0  # Increased from 2.0 to 5.0 seconds
            
            while time.time() - start_wait < max_wait:
                try:
                    section_tables = self.driver.find_elements(By.CLASS_NAME, "search-sectiontable")
                    term_headers = self.driver.find_elements(By.XPATH, "//h4[contains(text(), 'Spring') or contains(text(), 'Fall') or contains(text(), 'Summer')]")
                    
                    if len(section_tables) > 0 or len(term_headers) > 0:
                        return True
                    time.sleep(0.2)  # Check every 200ms
                except:
                    time.sleep(0.2)
            
            return True
            
        except Exception as e:
            print(f"❌ Failed to click section link: {e}")
            return False
    
    def extract_course_info(self, course_code: str, term: str) -> List[CourseSection]:
        """Extract detailed course information"""
        try:
            # Remove asterisk for course info extraction
            display_code = course_code.replace('*', ' ').strip()
            
            soup = BeautifulSoup(self.driver.page_source, 'html.parser')
            sections = []
            
            course_info = self.extract_basic_course_info(soup, display_code)
            
            term_headers = soup.find_all('h4', string=lambda t: t and any(term_word in t for term_word in ['Spring', 'Summer', 'Fall', 'Winter']) and '20' in t)
            
            # Filter to only the specified term if provided
            if term:
                term_headers = [h for h in term_headers if term in h.get_text(strip=True)]
            
            for term_header in term_headers:
                term_text = term_header.get_text(strip=True)
                term_sections = self.extract_sections_for_term(soup, term_header, term_text, course_info, course_code)
                sections.extend(term_sections)
            
            # Filter to keep FF sections OR sections with "hybrid" in locations
            filtered_sections = []
            for section in sections:
                # Keep if has FF in session code (and not WW)
                if "FF" in section.session_code and "WW" not in section.session_code:
                    filtered_sections.append(section)
                # Also keep if has "hybrid" in locations (case insensitive)
                elif any("hybrid" in location.lower() for location in section.locations):
                    filtered_sections.append(section)
            
            # Deduplicate sections based on course_code + session_code
            unique_sections = {}
            for section in filtered_sections:
                key = f"{section.course_code}_{section.session_code}"
                if key not in unique_sections:
                    unique_sections[key] = section
            
            return list(unique_sections.values())
            
        except Exception as e:
            print(f"❌ Course information extraction failed: {e}")
            return []
    
    def extract_basic_course_info(self, soup, course_name):
        """Extract basic course information"""
        course_info = {
            'title': f"{course_name} Course",
            'credits': '4',
            'description': '',
            'prerequisites': ''
        }
        
        # Clean course name for matching
        clean_name = course_name.replace('*', ' ').strip()
        parts = clean_name.split()
        if len(parts) >= 2:
            dept = parts[0].upper()
            num = parts[1]
            patterns = [
                lambda t: t and f"{dept}-{num}" in t.upper() and 'Credits' in t,
                lambda t: t and f"{dept} {num}" in t.upper() and 'Credits' in t,
                lambda t: t and f"{dept}*{num}" in t.upper() and 'Credits' in t,
                lambda t: t and dept in t.upper() and num in t and 'Credits' in t
            ]
            
            for pattern in patterns:
                title_elem = soup.find('span', string=pattern)
                if title_elem:
                    full_title = title_elem.get_text(strip=True)
                    # Clean the title to remove credit information
                    # Extract just the course code and title, remove credits part
                    if '(' in full_title and 'Credits' in full_title:
                        # Format: "DATA-610 Big Data Analytics/Data Mining (4 Credits)"
                        clean_title = full_title.split('(')[0].strip()
                        course_info['title'] = clean_title
                        # Extract credits number
                        credits_match = re.search(r'\((\d+)\s+Credits?\)', full_title)
                        if credits_match:
                            course_info['credits'] = credits_match.group(1)
                    else:
                        course_info['title'] = full_title
                    break
        
        return course_info
    
    def extract_sections_for_term(self, soup, term_header, term, course_info, course_name):
        """Extract all course sections for specific term"""
        sections = []
        
        current_element = term_header
        
        while current_element:
            current_element = current_element.find_next_sibling()
            if not current_element:
                break
                
            if current_element.name == 'h4' and any(t in current_element.get_text() for t in ['Spring', 'Summer', 'Fall', 'Winter']):
                break
            
            # Format course name for matching - ensure proper asterisk handling
            search_name = course_name.replace(' ', '*')
            if '*' not in search_name:
                parts = search_name.split('*')
                if len(parts) == 2:
                    search_name = f"{parts[0]}*{parts[1]}"
            
            # Find course section links - more flexible matching
            course_patterns = [
                search_name.upper(),
                course_name.upper().replace(' ', '*'),
                course_name.upper().replace(' ', '-'),
                course_name.upper().replace(' ', ''),
                course_name.split()[0].upper() if ' ' in course_name else course_name.upper()
            ]
            
            for pattern in course_patterns:
                section_links = current_element.find_all('a', string=lambda t: t and pattern in t.upper())
                if section_links:
                    for link in section_links:
                        section_info = self.extract_section_details(link, soup, term, course_info)
                        if section_info:
                            sections.append(section_info)
        
        return sections
    
    def extract_section_details(self, link_elem, soup, term, course_info) -> Optional[CourseSection]:
        """Extract section details"""
        try:
            full_section_code = link_elem.get_text(strip=True)
            
            # Parse section code: "DATA*610-Q1FF" -> course_code="DATA*610", session_code="Q1FF"
            if '-' in full_section_code:
                course_code, session_code = full_section_code.split('-', 1)
            else:
                course_code = full_section_code
                session_code = ""
            
            table = link_elem.find_parent('table')
            if not table:
                table = link_elem.find_next('table', class_='search-sectiontable')
            
            if not table:
                return None
            
            seats_info = self.extract_seats_info(table)
            time_info = self.extract_time_info(table)
            locations = self.extract_locations(table)
            instructor_info = self.extract_instructor_info(table)
            date_info = self.extract_date_info(table)
            
            section = CourseSection(
                course_code=course_code,
                session_code=session_code,
                course_name=course_info['title'],
                credits=course_info['credits'],
                seats_available=seats_info['available'],
                seats_total=seats_info['total'],
                seats_waitlisted=seats_info['waitlisted'],
                weekdays=time_info['weekdays'],  # Full weekday names
                weekdays_short=time_info['weekdays_short'],  # Short weekday names
                class_times=time_info['times'],
                locations=locations,
                instructors=instructor_info['instructors'],
                teaching_mode=instructor_info['mode'],
                start_date=date_info['start'],
                end_date=date_info['end'],
                term=term,
                prerequisites=course_info['prerequisites'],
                course_description=course_info['description']
            )
            
            return section
            
        except Exception as e:
            return None
    
    def extract_seats_info(self, table):
        """Extract seat information"""
        seats_info = {'available': 'N/A', 'total': 'N/A', 'waitlisted': 'N/A'}
        
        seat_span = table.find('span', class_='search-seatsavailabletext')
        if seat_span:
            seat_text = seat_span.get_text(strip=True)
            if '/' in seat_text:
                parts = seat_text.split('/')
                if len(parts) >= 3:
                    seats_info['available'] = parts[0].strip()
                    seats_info['total'] = parts[1].strip()
                    seats_info['waitlisted'] = parts[2].strip()
        
        return seats_info
    
    def extract_time_info(self, table):
        """Extract time information"""
        time_info = {
            'weekdays': [],
            'weekdays_short': [],
            'times': []
        }
        
        time_spans = table.find_all('span', class_='search-meetingtimestext')
        
        current_weekdays = []
        current_weekdays_short = []
        current_time = ""
        
        for span in time_spans:
            text = span.get_text(strip=True)
            
            # Check for weekday patterns - handle both single letters and "T/Th" format
            if (re.match(r'^[MTWRFSU]+\s*$', text) or 
                re.match(r'^(M|T|W|Th|TH|R|F|S|U)+\s*$', text) or
                re.match(r'^[MTWRFSU](/[MTWRFSU])*\s*$', text) or  # Handle T/Th, M/W/F, etc.
                re.match(r'^(T/Th|M/W|M/W/F|T/TH|Th/F)\s*$', text)):  # Common combinations
                # Save previous time block if exists
                if current_time and current_weekdays:
                    time_info['weekdays'].extend(current_weekdays)
                    time_info['weekdays_short'].extend(current_weekdays_short)
                    time_info['times'].append(current_time.strip())
                
                # Reset for new time block
                current_weekdays = []
                current_weekdays_short = []
                
                # Parse weekdays - handle different formats
                if '/' in text:
                    # Handle slash-separated format like "T/Th"
                    parts = text.split('/')
                    for part in parts:
                        part = part.strip()
                        if part in WEEKDAY_MAPPING:
                            current_weekdays.append(WEEKDAY_MAPPING[part])
                            current_weekdays_short.append(WEEKDAY_MAPPING_SHORT[part])
                else:
                    # Handle single letters format like "MTW"
                    i = 0
                    while i < len(text):
                        # Check for "Th" or "TH" first
                        if i < len(text) - 1 and text[i:i+2].upper() == 'TH':
                            current_weekdays.append('Thursday')
                            current_weekdays_short.append('Thu')
                            i += 2
                        # Single character weekdays
                        elif text[i] in WEEKDAY_MAPPING:
                            current_weekdays.append(WEEKDAY_MAPPING[text[i]])
                            current_weekdays_short.append(WEEKDAY_MAPPING_SHORT[text[i]])
                            i += 1
                        else:
                            i += 1
                
                current_time = ""
                
            elif re.match(r'\d{1,2}:\d{2}\s*(AM|PM)', text):
                current_time += text + " "
                
            elif text == '-':
                current_time += "- "
        
        # Add final time block
        if current_time and current_weekdays:
            time_info['weekdays'].extend(current_weekdays)
            time_info['weekdays_short'].extend(current_weekdays_short)
            time_info['times'].append(current_time.strip())
        
        # Remove duplicates while preserving order
        time_info['weekdays'] = list(dict.fromkeys(time_info['weekdays']))
        time_info['weekdays_short'] = list(dict.fromkeys(time_info['weekdays_short']))
        
        return time_info
    
    def extract_locations(self, table):
        """Extract class locations - include all location information"""
        locations = []
        
        location_cells = table.find_all('td', class_='search-sectionlocations')
        
        for cell in location_cells:
            # Extract all location information from spans
            spans = cell.find_all('span', class_='search-meetingtimestext')
            location_parts = []
            
            for span in spans:
                span_text = span.get_text(strip=True)
                
                # Only skip dates and truly empty content
                if (span_text and
                    span_text.upper() != 'TBD' and
                    not re.match(r'\d{1,2}/\d{1,2}/\d{4}', span_text) and  # Skip dates
                    span_text not in ['-', ',']):  # Skip separators
                    
                    # Split by commas and include all parts
                    parts = [p.strip() for p in span_text.split(',')]
                    for part in parts:
                        if part and part.upper() != 'TBD' and part not in ['-', ',']:
                            location_parts.append(part)
            
            if location_parts:
                location = ', '.join(location_parts)
                if location and location not in locations:
                    locations.append(location)
        
        return locations
    
    def extract_instructor_info(self, table):
        """Extract instructor information"""
        instructor_info = {
            'instructors': [],
            'mode': 'N/A'
        }
        
        instructor_cells = table.find_all('td', class_='search-sectioninstructormethods')
        
        for cell in instructor_cells:
            # Extract instructor names
            instructor_spans = cell.find_all('span', attrs={'aria-label': lambda x: x and 'Faculty Office Hours' in x})
            
            for span in instructor_spans:
                instructor_name = span.get_text(strip=True)
                if instructor_name and instructor_name not in instructor_info['instructors']:
                    instructor_info['instructors'].append(instructor_name)
            
            # Look for teaching mode in multiple ways:
            # 1. Look for bold spans with mode text
            mode_spans = cell.find_all('span', class_='bold')
            for span in mode_spans:
                mode_text = span.get_text(strip=True)
                if mode_text and ('Online' in mode_text or 'Blended' in mode_text or 'Face-To-Face' in mode_text or 'Face to Face' in mode_text or 'Hybrid' in mode_text):
                    instructor_info['mode'] = mode_text
                    break
            
            # 2. If no mode found, look in all spans within the cell
            if instructor_info['mode'] == 'N/A':
                all_spans = cell.find_all('span')
                for span in all_spans:
                    mode_text = span.get_text(strip=True)
                    if mode_text and ('Online' in mode_text or 'Blended' in mode_text or 'Face-To-Face' in mode_text or 'Face to Face' in mode_text or 'Hybrid' in mode_text):
                        instructor_info['mode'] = mode_text
                        break
            
            # 3. Look at parent elements after instructor spans
            if instructor_info['mode'] == 'N/A':
                for instructor_span in instructor_spans:
                    # Look at next siblings
                    next_element = instructor_span.find_next_sibling()
                    while next_element:
                        if hasattr(next_element, 'get_text'):
                            text = next_element.get_text(strip=True)
                            if text and ('Online' in text or 'Blended' in text or 'Face-To-Face' in text or 'Face to Face' in text or 'Hybrid' in text):
                                instructor_info['mode'] = text
                                break
                        next_element = next_element.find_next_sibling()
                    if instructor_info['mode'] != 'N/A':
                        break
                    
                    # Look at parent's next siblings
                    parent = instructor_span.parent
                    if parent:
                        next_parent = parent.find_next_sibling()
                        while next_parent:
                            if hasattr(next_parent, 'get_text'):
                                text = next_parent.get_text(strip=True)
                                if text and ('Online' in text or 'Blended' in text or 'Face-To-Face' in text or 'Face to Face' in text or 'Hybrid' in text):
                                    instructor_info['mode'] = text
                                    break
                            next_parent = next_parent.find_next_sibling()
                        if instructor_info['mode'] != 'N/A':
                            break
        
        # Fallback: Check location cells for mode information
        if instructor_info['mode'] == 'N/A':
            location_cells = table.find_all('td', class_='search-sectionlocations')
            for cell in location_cells:
                spans = cell.find_all('span', class_='search-meetingtimestext')
                for span in spans:
                    text = span.get_text(strip=True)
                    if 'Online' in text or 'Blended' in text or 'Face-To-Face' in text or 'Face to Face' in text or 'Hybrid' in text:
                        instructor_info['mode'] = text
                        break
                if instructor_info['mode'] != 'N/A':
                    break
        
        return instructor_info
    
    def extract_date_info(self, table):
        """Extract date information"""
        date_info = {'start': 'N/A', 'end': 'N/A'}
        
        date_spans = table.find_all('span', class_='search-meetingtimestext')
        
        for span in date_spans:
            text = span.get_text(strip=True)
            if re.match(r'\d{1,2}/\d{1,2}/\d{4}\s*-\s*\d{1,2}/\d{1,2}/\d{4}', text):
                dates = text.split('-')
                if len(dates) == 2:
                    date_info['start'] = dates[0].strip()
                    date_info['end'] = dates[1].strip()
                break
        
        return date_info
    
    def extract_classroom_number(self, locations):
        """Extract three-digit classroom number from location text"""
        if not locations:
            return ""
        
        # Join all location parts if it's a list
        location_text = ', '.join(locations) if isinstance(locations, list) else str(locations)
        
        # Look for three-digit numbers in the location text
        import re
        three_digit_pattern = r'\b\d{3}\b'
        matches = re.findall(three_digit_pattern, location_text)
        
        # Return the first three-digit number found, or empty string if none
        return matches[0] if matches else ""
    
    def save_to_excel(self, sections: List[CourseSection], filename: str = "on-campus_courses.xlsx"):
        """Save course information to Excel file with fresh sheet creation"""
        try:
            print(f"💾 Saving course information to {filename}...")
            
            # Generate sheet name with term and date
            term = sections[0].term if sections else "Unknown"
            today = datetime.now().strftime("%Y%m%d")
            sheet_name = f"{term}_{today}"
            # Clean sheet name (remove special characters and limit length)
            sheet_name = re.sub(r'[\\/*?:"<>|]', '', sheet_name)[:31]  # Excel limits sheet names to 31 chars
            
            # Check if file exists and preserve Note and Recitation data from existing sheet
            file_exists = os.path.exists(filename)
            existing_notes = {}  # key: course_code|session_code, value: {'note': '', 'recitation': ''}
            original_sheet_index = None  # Store original position
            
            if file_exists:
                try:
                    # Read existing data to preserve Note and Recitation columns
                    wb = load_workbook(filename)
                    if sheet_name in wb.sheetnames:
                        print(f"📄 Found existing sheet '{sheet_name}', preserving Note/Recitation data...")
                        
                        # Store the original position of the sheet
                        original_sheet_index = wb.index(wb[sheet_name])
                        print(f"📍 Original sheet position: {original_sheet_index}")
                        
                        # Read existing data with pandas
                        existing_df = pd.read_excel(filename, sheet_name=sheet_name)
                        
                        # Extract Note and Recitation data
                        if 'Course_Code' in existing_df.columns and 'Session_Code' in existing_df.columns:
                            for _, row in existing_df.iterrows():
                                key = f"{row.get('Course_Code', '')}|{row.get('Session_Code', '')}"
                                existing_notes[key] = {
                                    'note': row.get('Note', ''),
                                    'recitation': row.get('Recitation', '')
                                }
                        print(f"📝 Preserved data for {len(existing_notes)} existing sections")
                        
                        # Now remove the existing sheet
                        wb.remove(wb[sheet_name])
                        wb.save(filename)
                        print(f"🗑️  Removed existing sheet '{sheet_name}'")
                    
                    print(f"📄 Creating new sheet '{sheet_name}'...")
                except Exception as e:
                    print(f"⚠️  Could not preserve existing data: {e}")
            else:
                print(f"📄 Creating new file...")
            
            # Prepare new data
            new_data = []
            for section in sections:
                # Helper function to safely convert to int, return None for N/A
                def safe_int(value):
                    if value == 'N/A' or not value or not str(value).isdigit():
                        return None
                    try:
                        return int(value)
                    except:
                        return None
                
                # Calculate enrolled seats: Total - Available (as integer)
                enrolled_seats = None
                try:
                    if (section.seats_total != 'N/A' and section.seats_available != 'N/A' and 
                        section.seats_total.isdigit() and section.seats_available.isdigit()):
                        enrolled_seats = int(section.seats_total) - int(section.seats_available)
                except:
                    enrolled_seats = None
                
                # Get preserved Note and Recitation data for this section
                section_key = f"{section.course_code}|{section.session_code}"
                preserved_data = existing_notes.get(section_key, {'note': '', 'recitation': ''})
                
                # Convert classroom number to integer if it's a valid 3-digit number
                classroom_num = self.extract_classroom_number(section.locations)
                classroom_value = safe_int(classroom_num) if classroom_num else None
                
                new_data.append({
                    'Course_Code': section.course_code,
                    'Session_Code': section.session_code,
                    'Course_Name': section.course_name,
                    'Credits': safe_int(section.credits),  # Convert to int
                    'Enrolled_Seats': enrolled_seats,  # Already int or None
                    'Total_Seats': safe_int(section.seats_total),  # Convert to int
                    'Waitlist': safe_int(section.seats_waitlisted),  # Convert to int
                    'Weekdays': ', '.join(section.weekdays),
                    'Class_Times': ', '.join(section.class_times),
                    'Classroom': classroom_value,  # Convert to int
                    'Instructors': ', '.join(section.instructors),
                    'Teaching_Mode': section.teaching_mode,
                    'Start_Date': section.start_date,
                    'End_Date': section.end_date,
                    'Note': preserved_data['note'],  # Use preserved data
                    'First_Term': 'Yes' if section.is_first_term else 'No',  # String values
                    'Recitation': preserved_data['recitation']  # Use preserved data
                })
            
            new_df = pd.DataFrame(new_data)
            
            # Create new sheet at the original position
            if file_exists:
                # Use openpyxl directly for better position control
                wb = load_workbook(filename)
                if original_sheet_index is not None:
                    # Create sheet at the original position
                    new_ws = wb.create_sheet(sheet_name, original_sheet_index)
                    print(f"📍 Created sheet at original position {original_sheet_index}")
                else:
                    # Add to existing file (new sheet will be at the end)
                    new_ws = wb.create_sheet(sheet_name)
                
                # Write data to the worksheet
                # First write headers
                for c_idx, header in enumerate(new_df.columns, 1):
                    new_ws.cell(row=1, column=c_idx, value=header)
                
                # Then write all data rows
                for r_idx, row in enumerate(new_df.itertuples(index=False), 2):  # Start from row 2
                    for c_idx, value in enumerate(row, 1):
                        new_ws.cell(row=r_idx, column=c_idx, value=value)
                
                wb.save(filename)
            else:
                # Create new file
                with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                    new_df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            print(f"✅ Created new data in {filename} (Sheet: {sheet_name})")
            
            # Apply Excel formatting
            self.format_excel_table(filename, sheet_name)
                
        except Exception as e:
            print(f"❌ Failed to save Excel file: {e}")
    
    def format_excel_table(self, filename: str, sheet_name: str):
        """Format Excel sheet as a table with auto-adjusted column widths"""
        try:
            from openpyxl import load_workbook
            from openpyxl.worksheet.table import Table, TableStyleInfo
            from openpyxl.utils import get_column_letter
            
            # Load the workbook and worksheet
            wb = load_workbook(filename)
            ws = wb[sheet_name]
            
            # Remove existing tables from the worksheet to avoid conflicts
            tables_to_remove = list(ws.tables.keys())
            for table_name in tables_to_remove:
                del ws.tables[table_name]
            
            # Get the data range
            max_row = ws.max_row
            max_col = ws.max_column
            
            if max_row > 1:  # Ensure there's data beyond headers
                # Create table range (A1 to last cell with data)
                table_range = f"A1:{get_column_letter(max_col)}{max_row}"
                
                # Create a table with unique name - better cleaning for Excel table name requirements
                clean_sheet_name = sheet_name.replace(' ', '_').replace('-', '_').replace('/', '_').replace('(', '').replace(')', '')
                # Limit length and ensure it starts with a letter
                clean_sheet_name = clean_sheet_name[:20]  # Excel table names should be shorter
                table_name = f"CourseTable_{clean_sheet_name}"
                
                # Ensure table name is unique across the entire workbook
                counter = 1
                original_table_name = table_name
                existing_table_names = []
                for worksheet in wb.worksheets:
                    existing_table_names.extend([table.name for table in worksheet.tables.values()])
                
                while table_name in existing_table_names:
                    table_name = f"{original_table_name}_{counter}"
                    counter += 1
                
                # Ensure table name doesn't exceed Excel's limit (255 characters, but keep it shorter)
                if len(table_name) > 50:
                    table_name = table_name[:50]
                
                table = Table(displayName=table_name, ref=table_range)
                
                # Apply table style
                style = TableStyleInfo(
                    name="TableStyleMedium9",  # Blue table style
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=False
                )
                table.tableStyleInfo = style
                
                # Add the table to the worksheet
                ws.add_table(table)
                print(f"📊 Created table '{table_name}' with range {table_range}")
            
            # Auto-adjust column widths
            for column_cells in ws.columns:
                length = max(len(str(cell.value or "")) for cell in column_cells)
                # Set minimum width of 12 and maximum of 50
                adjusted_width = max(12, min(length + 2, 50))
                ws.column_dimensions[get_column_letter(column_cells[0].column)].width = adjusted_width
            
            # Freeze the first two columns
            ws.freeze_panes = 'C1'  # Freeze everything to the left of column C (i.e., freeze columns A and B)
            
            # Save the workbook
            wb.save(filename)
            print(f"📊 Applied Excel table formatting with auto-adjusted column widths and frozen first two columns")
            
            # Apply conditional formatting to compare with plan sheet
            self.apply_conditional_formatting(filename, sheet_name)
            
        except Exception as e:
            print(f"⚠️  Could not apply table formatting: {e}")
            import traceback
            print(f"⚠️  Error details: {traceback.format_exc()}")
            # Don't fail the entire operation if formatting fails
    
    def apply_conditional_formatting(self, filename: str, sheet_name: str):
        """
        Apply conditional formatting to highlight differences from plan sheet.
        
        Uses simplified formulas that properly handle None/NaN values for reliable
        change detection across all monitored columns. Applies highlighting to the
        PLAN sheet so users can see which planned values differ from current data.
        
        Args:
            filename (str): Excel file path
            sheet_name (str): Data sheet name to compare against
        """
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import PatternFill
            from openpyxl.formatting.rule import FormulaRule
            
            # Load workbook
            wb = load_workbook(filename)
            
            # Extract term and year from data sheet name
            # "Fall 2025_20250528" → term="Fall", year="2025"
            term, year = self.parse_sheet_name_for_comparison(sheet_name)
            if not term or not year:
                print(f"⚠️  Could not parse term/year from sheet name: {sheet_name}")
                return
            
            # Find corresponding plan sheet
            plan_sheet_name = f"{term} {year} Plan"
            if plan_sheet_name not in wb.sheetnames:
                print(f"📋 No plan sheet found: {plan_sheet_name}")
                return
            
            # Apply formatting to the PLAN sheet (not the data sheet)
            plan_ws = wb[plan_sheet_name]
            
            print(f"🔍 Applying conditional formatting to plan sheet: {plan_sheet_name}")
            print(f"🔍 Comparing against data sheet: {sheet_name}")
            
            # Define columns to compare with their names
            comparison_columns = {
                'E': 'Enrolled_Seats',
                'F': 'Total_Seats', 
                'G': 'Waitlist',
                'H': 'Weekdays',
                'I': 'Class_Times',
                'J': 'Classroom',
                'K': 'Instructors',
                'L': 'Teaching_Mode'
            }
            
            # Create yellow fill for differences
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            
            # Get data range from plan sheet (skip header row)
            max_row = plan_ws.max_row
            
            # Apply conditional formatting to highlight plan values that differ from data
            for col_letter, col_name in comparison_columns.items():
                try:
                    # Formula compares plan sheet cell against corresponding data sheet cell
                    # Highlights when plan value differs from current data
                    formula = f"={col_letter}2<>'{sheet_name}'!{col_letter}2"
                    
                    # Apply conditional formatting rule to plan sheet
                    rule = FormulaRule(formula=[formula], fill=yellow_fill)
                    range_to_format = f"{col_letter}2:{col_letter}{max_row}"
                    plan_ws.conditional_formatting.add(range_to_format, rule)
                    
                    print(f"  ✅ Added conditional formatting for {col_name} (column {col_letter})")
                    
                except Exception as e:
                    print(f"  ⚠️  Failed to add formatting for {col_name}: {e}")
            
            # Save workbook
            wb.save(filename)
            print(f"🎨 Applied conditional formatting to plan sheet '{plan_sheet_name}'")
            print(f"💡 Plan sheet will now highlight values that differ from current data")
            
        except Exception as e:
            print(f"⚠️  Could not apply conditional formatting: {e}")
            import traceback
            print(f"⚠️  Error details: {traceback.format_exc()}")
    
    def parse_sheet_name_for_comparison(self, sheet_name: str):
        """Parse sheet name to extract term and year"""
        try:
            # Handle formats like "Fall 2025_20250528"
            if '_' in sheet_name:
                base_name = sheet_name.split('_')[0]  # "Fall 2025"
            else:
                base_name = sheet_name  # "Fall 2025 Plan"
            
            # Split into term and year
            parts = base_name.strip().split()
            if len(parts) >= 2:
                term = parts[0]  # "Fall"
                year = parts[1]  # "2025" 
                return term, year
            
        except Exception as e:
            print(f"⚠️  Error parsing sheet name '{sheet_name}': {e}")
        
        return None, None
    
    def scrape_course(self, course_code: str, term: str) -> List[CourseSection]:
        """Complete course scraping workflow for single course"""
        try:
            print(f"🔍 Processing: {course_code} (Term: {term})")
            
            # Performance monitoring
            start_time = time.time()
            
            # Format course code consistently
            formatted_code = course_code
            if '*' not in formatted_code:
                parts = formatted_code.split()
                if len(parts) == 2:
                    formatted_code = f"{parts[0]}*{parts[1]}"
            
            if not self.search_course(formatted_code, term):
                return []
            
            search_time = time.time()
            
            if not self.click_view_sections(formatted_code):
                return []
            
            click_time = time.time()
            
            sections = self.extract_course_info(formatted_code, term)
            
            end_time = time.time()
            
            # Performance metrics
            search_duration = search_time - start_time
            click_duration = click_time - search_time
            extract_duration = end_time - click_time
            total_duration = end_time - start_time
            
            print(f"⏱️  Performance: Total={total_duration:.1f}s (Search={search_duration:.1f}s, Click={click_duration:.1f}s, Extract={extract_duration:.1f}s)")
            
            return sections
            
        except Exception as e:
            print(f"❌ Error processing {course_code}: {e}")
            return []
    
    def scrape_multiple_courses(self, course_request: CourseRequest) -> List[CourseSection]:
        """Batch scrape multiple courses for the same term"""
        all_sections = []
        
        print(f"📚 Processing {len(course_request.courses)} courses for term: {course_request.term}")
        
        # Create a map of course codes to their first_term status
        # Use multiple formats to handle format conversions between configuration and Franklin format
        first_term_map = {}
        for course_code, is_first_term in course_request.courses:
            # Add all possible variations of the course code
            variations = [
                course_code,  # Original: "DATA 610" 
                course_code.upper(),  # "DATA 610"
                course_code.replace(' ', '*'),  # "DATA*610"
                course_code.replace(' ', '*').upper(),  # "DATA*610"
                course_code.replace(' ', ''),  # "DATA610"
                course_code.replace(' ', '').upper(),  # "DATA610"
            ]
            
            for variation in variations:
                first_term_map[variation] = is_first_term
        
        for i, (course_code, is_first_term) in enumerate(course_request.courses):
            try:
                sections = self.scrape_course(course_code, course_request.term)
                if sections:
                    # Update First_Term for all sections of this course
                    for section in sections:
                        # Try multiple formats to find the first_term status
                        course_variations = [
                            section.course_code,  # e.g., "DATA*610"
                            section.course_code.replace('*', ' '),  # e.g., "DATA 610"
                            section.course_code.upper(),  # e.g., "DATA*610"
                            section.course_code.replace('*', ' ').upper()  # e.g., "DATA 610"
                        ]
                        
                        section.is_first_term = False  # Default
                        for variation in course_variations:
                            if variation in first_term_map:
                                section.is_first_term = first_term_map[variation]
                                break
                    all_sections.extend(sections)
                    print(f"✅ {course_code}: {len(sections)} sections found {'(First Term)' if is_first_term else ''}")
                else:
                    print(f"⚠️  {course_code}: No sections found")
                
                if i < len(course_request.courses) - 1:
                    time.sleep(1)
                    
            except Exception as e:
                print(f"❌ Failed to process {course_code}: {e}")
                continue
        
        return all_sections
    
    def close(self):
        """Close browser"""
        if self.driver:
            self.driver.quit()


def main():
    """Main function"""
    scraper = None
    try:
        print("🎯 Franklin University Course Scraper")
        print("=" * 50)
        
        scraper = FranklinCourseScraper(headless=False)
        
        course_request = scraper.read_course_list("course_request.md")
        
        if not course_request.courses:
            print("❌ No courses to process")
            return
        
        all_sections = scraper.scrape_multiple_courses(course_request)
        
        if all_sections:
            print(f"\n📊 Results: {len(all_sections)} sections from {len(course_request.courses)} courses (Term: {course_request.term})")
            
            # Save to Excel only
            scraper.save_to_excel(all_sections, "on-campus_courses.xlsx")
                
        else:
            print("❌ No course information found")
    
    except KeyboardInterrupt:
        print("\n⏹️  Operation interrupted")
    except Exception as e:
        print(f"❌ Program error: {e}")
    finally:
        if scraper:
            scraper.close()


if __name__ == "__main__":
    main() 